import logging
from functools import wraps
from datetime import timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from rest_framework import permissions

VALID_IMPORT_ACTIONS = {'baru', 'update'}
VALID_STOCK_MODES = {'set', 'tambah', 'kurangi'}

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - depends on Python environment
    Workbook = None
    load_workbook = None
from rest_framework.response import Response
from rest_framework.views import APIView

from .forms import CategoryForm, ItemForm, SupplierForm, TransactionForm
from .models import Category, Item, Supplier, Transaction

logger = logging.getLogger(__name__)


def get_user_role(user):
    if not user.is_authenticated:
        return 'guest'
    if user.is_superuser:
        return 'admin'

    admin_group, _ = Group.objects.get_or_create(name='Admin')
    operator_group, _ = Group.objects.get_or_create(name='Operator')

    if user.groups.filter(name='Admin').exists():
        return 'admin'
    if user.groups.filter(name='Operator').exists():
        return 'operator'

    user.groups.add(operator_group)
    return 'operator'


def role_required(*allowed_roles):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, 'Silakan login terlebih dahulu.')
                return redirect('login')

            role = get_user_role(request.user)
            if role not in allowed_roles:
                messages.error(request, 'Anda tidak memiliki izin untuk mengakses fitur ini.')
                return redirect('item_list')

            return view_func(request, *args, **kwargs)

        return _wrapped

    return decorator


def send_low_stock_notification(request, item, previous_stock=None):
    if item.stock > item.low_stock_threshold:
        return False

    if previous_stock is not None and previous_stock > item.low_stock_threshold:
        should_send = True
    else:
        should_send = previous_stock is None

    if not should_send:
        return False

    recipients = getattr(settings, 'LOW_STOCK_EMAIL_RECIPIENTS', [])
    if not recipients:
        recipients = [settings.DEFAULT_FROM_EMAIL]

    subject = f'Peringatan stok rendah: {item.name}'
    message = (
        f'Barang "{item.name}" ({item.code}) telah mencapai batas stok rendah.\n'
        f'Stok saat ini: {item.stock}.\n'
        f'Ambang batas: {item.low_stock_threshold}.'
    )

    try:
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, recipients)
        return True
    except Exception as exc:
        logger.exception('Failed to send low stock email for item %s', item.id)
        if request is not None:
            messages.warning(request, f'Notifikasi email gagal dikirim: {exc}')
        return False


@login_required
def item_list(request):
    """
    Menampilkan daftar barang dengan fitur pencarian,
    paginasi,
    dan ringkasan statistik aset.
    """
    query = request.GET.get('q', '')

    items = Item.objects.select_related('category', 'supplier').all().order_by('-created_at')

    if query:
        items = items.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(category__name__icontains=query)
            | Q(supplier__name__icontains=query)
        )

    paginator = Paginator(items, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_items = Item.objects.count()
    stats = Item.objects.aggregate(
        total_stock=Sum('stock'),
        total_value=Sum(F('stock') * F('price'))
    )
    low_stock_items = Item.objects.filter(stock__lte=F('low_stock_threshold')).order_by('stock')
    low_stock_count = low_stock_items.count()

    return render(request, 'inventory/item_list.html', {
        'page_obj': page_obj,
        'query': query,
        'total_items': total_items,
        'total_stock': stats['total_stock'] or 0,
        'total_value': stats['total_value'] or 0,
        'transaction_form': TransactionForm(),
        'low_stock_items': low_stock_items[:5],
        'low_stock_count': low_stock_count,
        'user_role': get_user_role(request.user),
    })


@login_required
@role_required('admin')
def item_create(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            if item.stock > 0:
                Transaction.objects.create(
                    item=item,
                    transaction_type='IN',
                    quantity=item.stock,
                    user=request.user,
                )
            if item.stock <= item.low_stock_threshold:
                send_low_stock_notification(request, item, None)

            messages.success(request, f"Barang '{item.name}' berhasil ditambahkan!")
            return redirect('item_list')
    else:
        form = ItemForm()
    return render(request, 'inventory/item_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def item_update(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        old_stock = item.stock
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            item = form.save()
            if item.stock != old_stock:
                quantity = abs(item.stock - old_stock)
                transaction_type = 'IN' if item.stock > old_stock else 'OUT'
                Transaction.objects.create(
                    item=item,
                    transaction_type=transaction_type,
                    quantity=quantity,
                    user=request.user,
                )
                if transaction_type == 'OUT' and item.stock <= item.low_stock_threshold:
                    send_low_stock_notification(request, item, old_stock)
            messages.info(request, f"Data '{item.name}' telah diperbarui.")
            return redirect('item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'inventory/item_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        nama_barang = item.name
        item.delete()
        messages.warning(request, f"Barang '{nama_barang}' telah dihapus dari sistem.")
        return redirect('item_list')
    return render(request, 'inventory/item_confirm_delete.html', {'item': item, 'user_role': get_user_role(request.user)})


@login_required
def item_detail(request, pk):
    item = get_object_or_404(Item, pk=pk)
    transactions = item.transactions.select_related('user').order_by('-timestamp')
    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/item_detail.html', {
        'item': item,
        'page_obj': page_obj,
        'transaction_form': TransactionForm(),
        'user_role': get_user_role(request.user),
    })


@login_required
@role_required('admin')
def category_list(request):
    categories = Category.objects.all().order_by('name')
    return render(request, 'inventory/category_list.html', {'categories': categories, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Kategori berhasil ditambahkan.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'inventory/category_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Kategori berhasil diperbarui.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'inventory/category_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.warning(request, 'Kategori berhasil dihapus.')
        return redirect('category_list')
    return render(request, 'inventory/category_confirm_delete.html', {'category': category, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by('name')
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier berhasil ditambahkan.')
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/supplier_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier berhasil diperbarui.')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/supplier_form.html', {'form': form, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.warning(request, 'Supplier berhasil dihapus.')
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_confirm_delete.html', {'supplier': supplier, 'user_role': get_user_role(request.user)})


@login_required
@role_required('admin')
def export_items_excel(request):
    if Workbook is None or load_workbook is None:
        messages.error(request, 'Fitur Excel membutuhkan paket openpyxl. Jalankan: pip install -r requirements.txt')
        return redirect('item_list')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Barang'

    # Header sesuai format import baru
    header = ['Aksi', 'Kode', 'Nama', 'Kategori', 'Supplier', 'Stok', 'Mode Stok', 'Harga', 'Ambang Stok Rendah']
    worksheet.append(header)

    # Tulis data item sesuai urutan header (aksi/mode dikosongkan untuk export data nyata)
    items = Item.objects.select_related('category', 'supplier').all().order_by('name')
    for item in items:
        row = [
            '',  # Aksi kosong saat export (user dapat memakai Template jika perlu contoh)
            item.code,
            item.name,
            item.category.name if item.category else '',
            item.supplier.name if item.supplier else '',
            int(item.stock),
            '',  # Mode Stok kosong saat export
            float(item.price) if item.price is not None else 0.0,
            int(item.low_stock_threshold),
        ]
        worksheet.append(row)
        # Set number formats for stok, harga, threshold
        r = worksheet.max_row
        worksheet.cell(row=r, column=6).number_format = '0'
        worksheet.cell(row=r, column=8).number_format = '#,##0.00'
        worksheet.cell(row=r, column=9).number_format = '0'

    # Buat sheet template terpisah dengan contoh agar data export tetap bersih
    template = workbook.create_sheet('TEMPLATE')
    template.append(header)
    template.append(['Baru', 'BRG001', 'Contoh Barang', 'Elektronik', 'PT Contoh', 20, 'Set', 150000, 5])
    template.append(['Update', 'BRG001', 'Contoh Barang Pro', 'Elektronik', 'PT Contoh', 5, 'Tambah', 180000, 7])
    template.column_dimensions['A'].width = 10
    template.column_dimensions['B'].width = 15
    template.column_dimensions['C'].width = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="items.xlsx"'
    workbook.save(response)
    return response


@login_required
@role_required('admin')
def import_items_excel(request):
    """
    Two-step import:
    1) Upload file -> parse and show preview (store preview in session)
    2) Confirm preview -> apply changes and create Transaction entries
    """
    if Workbook is None or load_workbook is None:
        messages.error(request, 'Fitur Excel membutuhkan paket openpyxl. Jalankan: pip install -r requirements.txt')
        return redirect('item_list')

    # Confirmation step
    if request.method == 'POST' and request.POST.get('confirm_import') == '1':
        preview = request.session.get('import_preview')
        if not preview:
            messages.error(request, 'Tidak ada data preview untuk diproses. Silakan unggah file Excel lagi.')
            return redirect('item_list')

        created_count = 0
        updated_count = 0
        transaction_count = 0

        for entry in preview:
            action = entry['action']
            code = entry['code']
            name = entry.get('name')
            category_name = entry.get('category')
            supplier_name = entry.get('supplier')
            price_value = entry.get('price')
            threshold_value = entry.get('threshold')
            stock_value = entry.get('stock')
            stock_mode = entry.get('stock_mode')
            delta = entry.get('delta')

            category, _ = Category.objects.get_or_create(name=category_name or 'Umum')
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name or 'Supplier Umum', defaults={'phone': '', 'email': ''})

            if action == 'baru':
                item = Item.objects.create(
                    code=code,
                    name=name or '',
                    category=category,
                    supplier=supplier,
                    stock=int(stock_value or 0),
                    price=price_value or 0,
                    low_stock_threshold=int(threshold_value or Item.LOW_STOCK_THRESHOLD),
                )
                created_count += 1
                # create transaction if stock > 0
                if int(stock_value or 0) > 0:
                    Transaction.objects.create(item=item, transaction_type='IN', quantity=int(stock_value or 0), user=request.user)
                    transaction_count += 1
            else:
                existing_item = Item.objects.filter(code=code).first()
                if existing_item is None:
                    continue
                old_stock = existing_item.stock
                if entry.get('name'):
                    existing_item.name = entry.get('name')
                existing_item.category = category
                existing_item.supplier = supplier
                if price_value is not None:
                    existing_item.price = price_value
                if threshold_value is not None:
                    existing_item.low_stock_threshold = int(threshold_value)

                # apply stock change according to preview
                if stock_value is not None:
                    if stock_mode == 'set':
                        existing_item.stock = int(stock_value)
                    elif stock_mode == 'tambah':
                        existing_item.stock += int(stock_value)
                    elif stock_mode == 'kurangi':
                        existing_item.stock = max(0, existing_item.stock - int(stock_value))

                existing_item.save()
                updated_count += 1

                # create transaction if delta != 0
                if delta is not None and int(delta) != 0:
                    if int(delta) > 0:
                        Transaction.objects.create(item=existing_item, transaction_type='IN', quantity=int(delta), user=request.user)
                    else:
                        Transaction.objects.create(item=existing_item, transaction_type='OUT', quantity=abs(int(delta)), user=request.user)
                    transaction_count += 1

        # clear preview
        try:
            del request.session['import_preview']
        except KeyError:
            pass

        messages.success(request, f'Import diterapkan. Dibuat: {created_count}, Diperbarui: {updated_count}, Transaksi dibuat: {transaction_count}.')
        return redirect('item_list')

    # Initial upload -> parse and show preview
    if request.method == 'POST' and request.FILES.get('excel_file'):
        workbook = load_workbook(request.FILES['excel_file'], data_only=True)
        worksheet = workbook.active

        headers = []
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]

        if not headers:
            messages.error(request, 'File Excel kosong.')
            return redirect('item_list')

        preview = []
        errors = []
        skipped = 0

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            row_data = dict(zip(headers, row))
            action = str(row_data.get('aksi') or row_data.get('action') or '').strip().lower()
            code = str(row_data.get('kode') or row_data.get('code') or '').strip()
            if not action or action not in VALID_IMPORT_ACTIONS:
                errors.append(f'Baris {row_number}: kolom Aksi harus berisi "Baru" atau "Update".')
                skipped += 1
                continue
            if not code:
                errors.append(f'Baris {row_number}: kode barang wajib diisi.')
                skipped += 1
                continue

            name = str(row_data.get('nama') or row_data.get('name') or '').strip()
            if action == 'baru' and not name:
                errors.append(f'Baris {row_number}: nama barang wajib diisi untuk aksi Baru.')
                skipped += 1
                continue

            category_name = str(row_data.get('kategori') or row_data.get('category') or '').strip()
            supplier_name = str(row_data.get('supplier') or row_data.get('supplier_name') or '').strip()
            stock_value = row_data.get('stok') or row_data.get('stock') or None
            price_value = row_data.get('harga') or row_data.get('price') or None
            threshold_value = row_data.get('ambang stok rendah') or row_data.get('low_stock_threshold') or None
            stock_mode = str(row_data.get('mode stok') or row_data.get('stock_mode') or 'set').strip().lower()
            if stock_mode not in VALID_STOCK_MODES:
                errors.append(f'Baris {row_number}: mode stok tidak valid. Gunakan Set, Tambah, atau Kurangi.')
                skipped += 1
                continue

            existing_item = Item.objects.filter(code=code).first()
            if action == 'baru' and existing_item is not None:
                errors.append(f'Baris {row_number}: kode {code} sudah ada. Gunakan aksi Update.')
                skipped += 1
                continue
            if action == 'update' and existing_item is None:
                errors.append(f'Baris {row_number}: kode {code} tidak ditemukan. Gunakan aksi Baru.')
                skipped += 1
                continue

            # compute delta for preview
            old_stock = existing_item.stock if existing_item is not None else 0
            new_stock = None
            delta = None
            if action == 'baru':
                new_stock = int(stock_value or 0)
                delta = new_stock
            else:
                if stock_value is None:
                    new_stock = old_stock
                    delta = 0
                else:
                    sv = int(stock_value)
                    if stock_mode == 'set':
                        new_stock = sv
                    elif stock_mode == 'tambah':
                        new_stock = old_stock + sv
                    else:
                        new_stock = max(0, old_stock - sv)
                    delta = new_stock - old_stock

            preview.append({
                'row': row_number,
                'action': action,
                'code': code,
                'name': name,
                'category': category_name,
                'supplier': supplier_name,
                'stock': int(stock_value) if stock_value is not None else None,
                'stock_mode': stock_mode,
                'price': float(price_value) if price_value is not None else None,
                'threshold': int(threshold_value) if threshold_value is not None else None,
                'old_stock': old_stock,
                'new_stock': new_stock,
                'delta': delta,
            })

        # store preview in session
        request.session['import_preview'] = preview
        request.session.modified = True

        # render preview page
        return render(request, 'inventory/import_preview.html', {
            'preview': preview,
            'errors': errors,
            'skipped': skipped,
        })

    return redirect('item_list')


@login_required
def add_transaction(request, item_id):
    """
    Menangani mutasi stok (IN/OUT) dengan validasi ketersediaan stok.
    """
    item = get_object_or_404(Item, id=item_id)

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.item = item
            transaction.user = request.user

            previous_stock = item.stock
            if transaction.transaction_type == 'IN':
                item.stock += transaction.quantity
                item.save()
                transaction.save()
                messages.success(request, f"Stok {item.name} berhasil ditambah!")
            elif transaction.transaction_type == 'OUT':
                if item.stock >= transaction.quantity:
                    item.stock -= transaction.quantity
                    item.save()
                    transaction.save()
                    messages.success(request, f"Stok {item.name} berhasil dikurangi!")
                    if item.stock <= item.low_stock_threshold:
                        send_low_stock_notification(request, item, previous_stock)
                else:
                    messages.error(request, f"Gagal! Stok {item.name} tidak mencukupi.")

            redirect_to = request.POST.get('next')
            if redirect_to and redirect_to.startswith('/'):
                return redirect(redirect_to)
            return redirect('item_list')

    return redirect('item_list')


@login_required
def transaction_history(request):
    """
    Menampilkan riwayat mutasi dengan filter tanggal.
    """
    transactions = Transaction.objects.select_related('item', 'user').all().order_by('-timestamp')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        transactions = transactions.filter(timestamp__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(timestamp__date__lte=end_date)

    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/transaction_history.html', {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'user_role': get_user_role(request.user),
    })


@login_required
def export_transactions_csv(request):
    import csv

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="laporan_mutasi.csv"'

    writer = csv.writer(response)
    writer.writerow(['Waktu', 'Petugas', 'Barang', 'Tipe', 'Jumlah'])

    transactions = Transaction.objects.all().order_by('-timestamp')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and start_date != "":
        transactions = transactions.filter(timestamp__date__gte=start_date)
    if end_date and end_date != "":
        transactions = transactions.filter(timestamp__date__lte=end_date)

    for transaction in transactions:
        writer.writerow([
            transaction.timestamp.strftime('%Y-%m-%d %H:%M'),
            transaction.user.username if transaction.user else 'System',
            transaction.item.name,
            transaction.get_transaction_type_display(),
            transaction.quantity,
        ])

    return response


class InventoryChartData(APIView):
    """
    Endpoint untuk menyediakan data JSON bagi Chart.js di frontend.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, format=None):
        today = timezone.now().date()
        labels = []
        data_in = []
        data_out = []

        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime('%d %b'))

            in_qty = Transaction.objects.filter(
                timestamp__date=date,
                transaction_type='IN'
            ).aggregate(total=Sum('quantity'))['total'] or 0

            out_qty = Transaction.objects.filter(
                timestamp__date=date,
                transaction_type='OUT'
            ).aggregate(total=Sum('quantity'))['total'] or 0

            data_in.append(in_qty)
            data_out.append(out_qty)

        return Response({
            'labels': labels,
            'datasets': [
                {
                    'label': 'Masuk',
                    'data': data_in,
                    'borderColor': '#198754',
                    'backgroundColor': 'rgba(25, 135, 84, 0.1)',
                    'tension': 0.3,
                    'fill': True,
                },
                {
                    'label': 'Keluar',
                    'data': data_out,
                    'borderColor': '#dc3545',
                    'backgroundColor': 'rgba(220, 53, 69, 0.1)',
                    'tension': 0.3,
                    'fill': True,
                },
            ],
        })
